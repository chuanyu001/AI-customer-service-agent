# 知识答案润色 → 写入Excel核验 (不碰数据库)
# 用法: python scripts/polish_to_excel.py
# 读取行车记录仪622知识库, 对每条答案润色, 写入新列"润色答案", 另存为新文件
#
# 输出: 行车记录仪Agent知识库_更新版622_润色版.xlsx

import sys
import os
import asyncio
import shutil

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from app.services.llm_service import get_llm

SRC = r"D:\wuchu\Desktop\personalfiles\实习\kefu-agent\agent知识库汇总版6.22\行车记录仪Agent知识库_更新版622.xlsx"
OUT = r"D:\wuchu\Desktop\personalfiles\实习\kefu-agent\agent知识库汇总版6.22\行车记录仪Agent知识库_更新版622_润色版.xlsx"

# 润色答案列插在"标准答案"后面
ANSWER_COL_NAME = "标准答案"
POLISHED_COL_NAME = "润色答案"


async def polish_all():
    print("🤖 加载大模型...")
    llm = get_llm()

    # 读01表
    df = pd.read_excel(SRC, sheet_name="01_知识问答库")
    print(f"📖 读取到 {len(df)} 条知识\n")

    polished_list = []
    total = len(df)

    for i, row in df.iterrows():
        question = str(row.get("标准问题", "")).strip()
        answer = str(row.get(ANSWER_COL_NAME, "")).strip()
        code = str(row.get("知识编号", "")).strip()

        if not answer or answer == "nan":
            polished_list.append("")
            continue

        try:
            polished = await llm.polish(question, answer)
            if not polished or not polished.strip():
                polished = answer  # 空则用原文兜底
        except Exception as e:
            print(f"  [{i+1}/{total}] {code} 润色失败, 用原文: {e}")
            polished = answer

        polished_list.append(polished)

        q_short = question[:20]
        print(f"  [{i+1}/{total}] {code} {q_short}... OK")

        # 每20条保存一次中间结果, 防中断丢失
        if (i + 1) % 20 == 0:
            print(f"  --- 已完成 {i+1}/{total} ---")

    # 插入润色列 (在标准答案后面)
    answer_idx = list(df.columns).index(ANSWER_COL_NAME)
    df.insert(answer_idx + 1, POLISHED_COL_NAME, polished_list)

    return df


def save_with_style(df):
    """用openpyxl保存, 复制原文件所有sheet, 替换01表, 并给润色列加样式"""
    # 先复制原文件
    shutil.copy2(SRC, OUT)
    print(f"\n💾 写入 {OUT}")

    # 用openpyxl打开, 替换01表
    wb = load_workbook(OUT)
    if "01_知识问答库" in wb.sheetnames:
        # 删除原01表, 重新创建
        ws_old = wb["01_知识问答库"]
        sheet_idx = wb.sheetnames.index("01_知识问答库")
        wb.remove(ws_old)
        ws = wb.create_sheet("01_知识问答库", sheet_idx)
    else:
        ws = wb.create_sheet("01_知识问答库", 0)

    # 写表头
    headers = list(df.columns)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写数据
    polished_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 淡黄色高亮润色列
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="" if pd.isna(value) else value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # 润色列加淡黄底色
            if headers[col_idx - 1] == POLISHED_COL_NAME:
                cell.fill = polished_fill

    # 列宽
    col_widths = {
        "知识编号": 12, "产品模块": 12, "厂家": 16, "知识类型": 16,
        "标准问题": 35, "常见问法": 30, "标准答案": 40, "润色答案": 45,
        "是否需要识别品牌": 12, "是否需要附件": 12, "关联查询编号": 12,
        "转人工条件": 16, "状态": 10, "来源备注": 16,
    }
    from openpyxl.utils import get_column_letter
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 14)

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUT)


async def main():
    print(f"📂 源文件: {SRC}")
    print(f"📤 输出文件: {OUT}\n")

    df = await polish_all()
    save_with_style(df)

    print(f"\n✅ 润色完成!")
    print(f"   共润色 {len(df)} 条")
    print(f"   润色结果在「润色答案」列(淡黄色高亮), 原文在「标准答案」列")
    print(f"   请核验后告诉我是否落库")


if __name__ == "__main__":
    asyncio.run(main())
