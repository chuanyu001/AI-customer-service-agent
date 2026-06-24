# -*- coding: utf-8 -*-
"""为三个业务知识库 6.24 润色版套用记录仪知识库的 Excel 样式 (增强可读性)。

记录仪样式:
  表头: 深蓝底(#2F5496) + 白字加粗 + 水平/垂直居中, Carlito 11pt
  数据行: 顶对齐 + 自动换行(wrap), Carlito 11pt 非加粗
  各列固定列宽, 冻结表头行
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r"D:/wuchu/Desktop/personalfiles/实习/kefu-agent/agent知识库汇总版6.22"
FILES = [
    "折扣加油知识库润色版6.24.xlsx",
    "WiFi套餐知识库润色版6.24.xlsx",
    "基础流量处理知识库润色版6.24.xlsx",
]

KNOWLEDGE_SHEET = "01_知识问答库"

# 列宽 (A-M, 与记录仪一致)
COL_WIDTHS = {
    "A": 12.0,  # 知识编号
    "B": 13.0,  # 产品模块
    "C": 16.0,  # 厂家
    "D": 13.0,  # 知识类型
    "E": 35.0,  # 标准问题
    "F": 30.0,  # 常见问法
    "G": 40.0,  # 标准答案
    "H": 45.0,  # 润色答案
    "I": 25.5,  # 是否需要识别品牌
    "J": 30.0,  # 是否需要附件
    "K": 12.0,  # 关联查询编号
    "L": 16.0,  # 状态
    "M": 10.0,  # 答复策略
}

HEADER_FONT = Font(name="Carlito", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Carlito", size=11, bold=False)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
# 细边框增强表格可读性
THIN = Side(style="thin", color="FFBFBFBF")
DATA_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USAGE_TITLE_FONT = Font(name="Carlito", size=14, bold=True, color="FF2F5496")
USAGE_HEADER_FONT = Font(name="Carlito", size=11, bold=True, color="FFFFFFFF")
USAGE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF2F5496")


def style_sheet(ws):
    max_col = ws.max_column
    max_row = ws.max_row

    # 列宽
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 表头行
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = DATA_BORDER
    ws.row_dimensions[1].height = 28

    # 数据行
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = DATA_BORDER
        # 行高留空, Excel 会根据 wrap_text 自动适应

    # 冻结表头行
    ws.freeze_panes = "A2"


def style_usage_sheet(ws):
    """00_使用说明: 标题行蓝色加粗, 表头行蓝底白字"""
    max_row = ws.max_row
    max_col = ws.max_column
    # 第1行作为标题
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = USAGE_TITLE_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24
    # 若有第2行表头, 加蓝底
    if max_row >= 2:
        for c in range(1, max_col + 1):
            cell = ws.cell(row=2, column=c)
            cell.font = USAGE_HEADER_FONT
            cell.fill = USAGE_HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = DATA_BORDER
        ws.row_dimensions[2].height = 22
    # 其余数据行换行顶对齐
    for r in range(3, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = DATA_BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def main():
    for f in FILES:
        path = os.path.join(BASE, f)
        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            if sheet_name == KNOWLEDGE_SHEET:
                style_sheet(wb[sheet_name])
            elif sheet_name == "00_使用说明":
                style_usage_sheet(wb[sheet_name])
        wb.save(path)
        print(f"✅ {f}: 样式已套用 (表头蓝底白字/数据行换行顶对齐/列宽/冻结表头)")


if __name__ == "__main__":
    main()
