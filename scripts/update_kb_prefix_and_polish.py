# 更新知识编号前缀 KB→JL, 并导入润色答案
# 1. Excel: KB0001 → JL0001 (润色版文件)
# 2. 数据库: knowledge_answer.knowledge_code 批量改 KB→JL
#           knowledge_attachment 等关联表无需改(用knowledge_id数字主键)
# 3. 数据库: 导入润色答案到 polished_answer 字段
#
# 用法: python scripts/update_kb_prefix_and_polish.py

import sys
import os
import asyncio
import re

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import pandas as pd
from sqlalchemy import select, update
from openpyxl import load_workbook
from app.core.database import get_session_factory
from app.models import KnowledgeAnswer

EXCEL_PATH = r"D:\wuchu\Desktop\personalfiles\实习\kefu-agent\agent知识库汇总版6.22\行车记录仪Agent知识库_更新版622_润色版.xlsx"
SHEET_NAME = "01_知识问答库"


def update_excel_prefix():
    """修改Excel中知识编号前缀 KB→JL"""
    print("📝 修改Excel编号前缀 KB→JL...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 找"知识编号"列
    header_row = [c.value for c in ws[1]]
    if "知识编号" not in header_row:
        print("  ⚠️ 未找到'知识编号'列")
        return 0
    code_col = header_row.index("知识编号") + 1

    count = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=code_col)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("KB"):
            cell.value = "JL" + cell.value[2:]
            count += 1

    wb.save(EXCEL_PATH)
    print(f"  ✅ Excel已修改 {count} 个编号: KB→JL")
    return count


async def update_database_and_import_polish():
    """数据库: 改编号前缀 + 导入润色答案"""
    print("\n📦 读取Excel(已改前缀)...")
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)

    # 构建映射: 旧编号(KB) → {新编号(JL), 润色答案}
    updates = {}
    for _, row in df.iterrows():
        code = str(row.get("知识编号", "")).strip()
        polished = str(row.get("润色答案", "")).strip()
        if code.startswith("JL"):
            old_code = "KB" + code[2:]
            updates[old_code] = {"new_code": code, "polished": polished}

    print(f"   读取到 {len(updates)} 条待更新\n")

    async with get_session_factory()() as db:
        # Step 1: 改编号前缀 KB→JL
        print("🔄 更新数据库编号前缀 KB→JL...")
        prefix_count = 0
        polish_count = 0
        for old_code, info in updates.items():
            stmt = select(KnowledgeAnswer).where(KnowledgeAnswer.knowledge_code == old_code)
            result = await db.execute(stmt)
            ka = result.scalar_one_or_none()
            if ka:
                ka.knowledge_code = info["new_code"]
                # 同时导入润色答案
                if info["polished"] and info["polished"] != "nan":
                    ka.polished_answer = info["polished"]
                    polish_count += 1
                prefix_count += 1
        await db.commit()

    print(f"  ✅ 编号前缀更新: {prefix_count} 条")
    print(f"  ✅ 润色答案导入: {polish_count} 条")


async def main():
    print(f"📂 Excel文件: {EXCEL_PATH}\n")

    # 1. 改Excel前缀
    update_excel_prefix()

    # 2&3. 改数据库编号 + 导入润色答案
    await update_database_and_import_polish()

    print("\n✅ 全部完成!")
    print("   - Excel编号: KB→JL")
    print("   - 数据库编号: KB→JL")
    print("   - 润色答案: 已导入 polished_answer 字段")


if __name__ == "__main__":
    asyncio.run(main())
