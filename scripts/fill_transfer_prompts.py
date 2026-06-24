# -*- coding: utf-8 -*-
"""为四张知识库里「是否对客=转人工」的条目, 在「答复策略」列填入转人工前追问语。
对客类条目不动。追问语用于 AI 命中转人工类知识时先追问关键信息, 收齐再转人工。
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

BASE = r"D:/wuchu/Desktop/personalfiles/实习/kefu-agent/agent知识库汇总版"

# 转人工前追问语 (知识编号 → 追问语)
PROMPTS = {
    # 折扣加油
    "JY0006": "请问您是企业车队用户、需要车队加油开票吗？方便留个联系电话，我帮您对接商务同事处理。",
    # WiFi套餐
    "WF0008": "请问您的车型是鹰途/JH6，还是其他车系？方便提供一下车牌号吗，我帮您核实并转人工处理。",
    # 基础流量 — A类: 需收车辆身份(VIN/ICCID)
    "LL0002": "请问您车辆的车架号（VIN）或ICCID号是多少？我帮您转人工客服核实处理。",
    "LL0004": "请问您车辆的车架号（VIN）或ICCID号是多少？我帮您转人工客服核实处理。",
    "LL0010": "请问您车辆的车架号（VIN）或ICCID号是多少？我帮您转人工客服核实处理。",
    "LL0011": "请问您车辆的车架号（VIN）或ICCID号是多少？我帮您转人工客服核实处理。",
    "LL0012": "请问您车辆的车架号（VIN）或ICCID号是多少？我帮您转人工客服核实处理。",
    "LL0014": "请问您车辆的车架号（VIN）或ICCID号是多少？我帮您转人工客服核实处理。",
    "LL0017": "请问您车辆的车架号（VIN）或ICCID号是多少？我帮您转人工客服核实处理。",
    "LL0018": "请问您车辆的车架号（VIN）或ICCID号是多少？我帮您转人工客服核实处理。",
    # 基础流量 — B类: 概念/流程咨询类
    "LL0005": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0006": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0007": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0008": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0009": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0013": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0015": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0016": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0019": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
    "LL0020": "这个问题需要人工客服按内部流程为您处理，请问方便提供车架号（VIN）和您遇到的具体情况吗？我帮您转接。",
}

DATA_FONT = Font(name="Carlito", size=11, bold=False)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILES = [
    "记录仪知识库润色版6.24.xlsx",
    "折扣加油知识库润色版6.24.xlsx",
    "WiFi套餐知识库润色版6.24.xlsx",
    "基础流量处理知识库润色版6.24.xlsx",
]


def process(filename):
    path = os.path.join(BASE, filename)
    wb = load_workbook(path)
    ws = wb["01_知识问答库"]

    # 定位列
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    code_col = headers["知识编号"]
    flag_col = headers["是否对客"]
    strategy_col = headers["答复策略"]

    filled = 0
    skipped = []
    for r in range(2, ws.max_row + 1):
        flag = ws.cell(row=r, column=flag_col).value
        if flag != "转人工":
            continue
        code = str(ws.cell(row=r, column=code_col).value).strip()
        prompt = PROMPTS.get(code)
        if not prompt:
            skipped.append(code)
            continue
        cell = ws.cell(row=r, column=strategy_col)
        cell.value = prompt
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = BORDER
        filled += 1

    wb.save(path)
    return filled, skipped


def main():
    for f in FILES:
        filled, skipped = process(f)
        msg = f"，无追问语配置: {skipped}" if skipped else ""
        print(f"✅ {f}: 填入追问语 {filled} 条{msg}")


if __name__ == "__main__":
    main()
