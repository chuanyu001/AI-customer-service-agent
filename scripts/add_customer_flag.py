# -*- coding: utf-8 -*-
"""为四张知识库(记录仪 + 折扣加油 + WiFi套餐 + 基础流量处理)新增「是否对客」列。
取值: 对客 / 转人工 / 内部
  - 对客  : AI 可直接自动回复
  - 转人工: 命中即转人工(需核实客户信息/工单类/商务跟进)
  - 内部  : 客服坐席内部SOP, 不进对客检索库, 命中转人工

记录仪库参照「答复策略」字段辅助判断; 三个业务库按逐条审核结果标注。
列插入位置: 「状态」列之前 (与答复策略等控制类字段相邻)。
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r"D:/wuchu/Desktop/personalfiles/实习/kefu-agent/agent知识库汇总版6.22"

# ========== 三个业务库逐条审核结果 (知识编号 → 是否对客) ==========
# 依据上一轮审读结论
BUSINESS_CUSTOMER_FLAG = {
    # 折扣加油: 7条可对客, JY0006 含商务跟进 → 转人工
    "JY0001": "对客", "JY0002": "对客", "JY0003": "对客", "JY0004": "对客",
    "JY0005": "对客", "JY0006": "转人工", "JY0007": "对客", "JY0008": "对客",
    # WiFi: 7条对客, WF0008 内部SOP → 转人工
    "WF0001": "对客", "WF0002": "对客", "WF0003": "对客", "WF0004": "对客",
    "WF0005": "对客", "WF0006": "对客", "WF0007": "对客",
    "WF0008": "转人工", "WF0009": "对客",
    # 基础流量: 仅LL0001对客, 其余内部SOP → 内部/转人工
    "LL0001": "对客",
    # LL0002 青岛换2.0需核实处理 → 转人工
    "LL0002": "转人工",
    # LL0003 纯概念区分可对客? 含"反馈和充值流程"内部术语 → 对客(产品概念部分), 但提及运营确认 → 标对客(知识本身是区分方法, 可告知客户)
    "LL0003": "对客",
    # LL0004-0020 全是内部反馈/充值/登记SOP → 内部
    "LL0004": "内部", "LL0005": "内部", "LL0006": "内部", "LL0007": "内部",
    "LL0008": "内部", "LL0009": "内部", "LL0010": "内部", "LL0011": "内部",
    "LL0012": "内部", "LL0013": "内部", "LL0014": "内部", "LL0015": "内部",
    "LL0016": "内部", "LL0017": "内部", "LL0018": "内部", "LL0019": "内部",
    "LL0020": "内部",
}

# 记录仪库: 默认对客; 答复策略提示需交互排查/无标准答案/重复 的 → 转人工
DASHCAM_TRANSFER_KEYWORDS = ["无标准答案", "需交互排查", "重复"]


def classify_dashcam(row) -> str:
    """记录仪库分类规则"""
    strategy = str(row.get("答复策略", "")) if pd.notna(row.get("答复策略")) else ""
    # 答复策略含交互排查/无标准答案/重复 → 转人工
    if any(kw in strategy for kw in DASHCAM_TRANSFER_KEYWORDS):
        return "转人工"
    return "对客"


# 注: 仅两种选项 对客/转人工, 内部SOP 命中也按转人工处理
INTERNAL_AS = "转人工"

# 列插入位置: 「状态」列之前
INSERT_BEFORE = "状态"
NEW_COL = "是否对客"

# 样式 (与记录仪库一致)
HEADER_FONT = Font(name="Carlito", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Carlito", size=11, bold=False)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True, horizontal="center")
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
# 不同标记用不同底色, 增强可读性
FLAG_FILL = {
    "对客": PatternFill(fill_type="solid", fgColor="FFE2EFDA"),   # 浅绿
    "转人工": PatternFill(fill_type="solid", fgColor="FFFFF2CC"), # 浅黄
    "内部": PatternFill(fill_type="solid", fgColor="FFFCE4D6"),   # 浅橙
}


def get_flag(df: pd.DataFrame) -> pd.Series:
    """生成「是否对客」标记列"""
    if "知识编号" in df.columns and df["知识编号"].iloc[0] if len(df) else "":
        first_code = str(df["知识编号"].iloc[0]) if len(df) else ""
        if first_code.startswith("JY") or first_code.startswith("WF") or first_code.startswith("LL"):
            # 三个业务库
            return df["知识编号"].map(lambda c: BUSINESS_CUSTOMER_FLAG.get(str(c).strip(), "对客"))
        elif first_code.startswith("KB") or first_code.startswith("JL"):
            # 记录仪库 (知识编号前缀以 KB/JL 等开头, 或其他) → 用分类规则
            return df.apply(classify_dashcam, axis=1)
    # 兜底: 默认对客
    return pd.Series(["对客"] * len(df))


def add_column_to_file(filename: str):
    path = os.path.join(BASE, filename)
    wb = load_workbook(path)
    ws = wb["01_知识问答库"]

    # 定位列
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    if NEW_COL in headers:
        print(f"  ⚠ {filename} 已存在「{NEW_COL}」列, 跳过")
        return
    status_col = headers.get(INSERT_BEFORE)
    if status_col is None:
        raise RuntimeError(f"{filename} 未找到「{INSERT_BEFORE}」列")

    insert_at = status_col
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = NEW_COL

    # 读数据生成标记
    import pandas as pd
    df = pd.read_excel(path, sheet_name="01_知识问答库")
    flags = get_flag(df)

    # 填充 + 样式
    counts = {}
    for i, flag in enumerate(flags):
        r = i + 2
        cell = ws.cell(row=r, column=insert_at)
        cell.value = flag
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = BORDER
        cell.fill = FLAG_FILL.get(flag, FLAG_FILL["对客"])
        counts[flag] = counts.get(flag, 0) + 1

    # 表头样式
    hcell = ws.cell(row=1, column=insert_at)
    hcell.font = HEADER_FONT
    hcell.fill = HEADER_FILL
    hcell.alignment = HEADER_ALIGN
    hcell.border = BORDER
    # 新增列宽
    ws.column_dimensions[hcell.column_letter].width = 12

    wb.save(path)
    print(f"✅ {filename}: 新增「{NEW_COL}」列 → {counts}")


def main():
    for f in [
        "记录仪知识库润色版6.24.xlsx",
        "折扣加油知识库润色版6.24.xlsx",
        "WiFi套餐知识库润色版6.24.xlsx",
        "基础流量处理知识库润色版6.24.xlsx",
    ]:
        add_column_to_file(f)


if __name__ == "__main__":
    main()
