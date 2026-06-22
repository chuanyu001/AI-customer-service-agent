# 知识库汇总表样式优化脚本
# 参照行车记录仪6.22原表样式, 给"知识库汇总版6.22.xlsx"套用统一样式
#
# 样式特征 (来自原表):
# - 表头: 深蓝填充2F75B5, 白字加粗, 居中, 自动换行, 行高34
# - 数据行: Carlito 11, 顶部对齐, 自动换行
# - 列宽: 按列内容设置
# - 冻结首行
#
# 用法: python scripts/style_knowledge_base.py

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = r"D:\wuchu\Desktop\personalfiles\实习\kefu-agent\agent知识库汇总版6.22\知识库汇总版6.22.xlsx"

# 表头样式
HEADER_FONT = Font(name="Carlito", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(start_color="FF2F75B5", end_color="FF2F75B5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 数据样式
DATA_FONT = Font(name="Carlito", size=11)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)

# 边框 (浅灰, 让表格更清晰)
THIN_BORDER = Border(
    left=Side(style="thin", color="FFD9D9D9"),
    right=Side(style="thin", color="FFD9D9D9"),
    top=Side(style="thin", color="FFD9D9D9"),
    bottom=Side(style="thin", color="FFD9D9D9"),
)

# 各列推荐宽度 (按列名, 未命中的给默认值)
COLUMN_WIDTHS = {
    "知识编号": 11,
    "产品模块": 12,
    "厂家": 14,
    "知识类型": 18,
    "标准问题": 34,
    "常见问法": 40,
    "标准答案": 48,
    "是否需要识别品牌": 18,
    "是否需要附件": 14,
    "关联查询编号": 28,
    "转人工条件": 18,
    "状态": 12,
    "来源备注": 24,
    # 查询问题库
    "查询编号": 12,
    "查询意图": 22,
    "用户必填信息": 20,
    "查询数据源": 16,
    "匹配条件": 28,
    "返回字段": 28,
    "正常回复模板": 40,
    "空值/未匹配回复模板": 30,
    "转人工规则": 18,
    "关联知识类型": 16,
    "是否允许自动回复": 14,
    "备注": 18,
    # 品牌识别规则
    "优先级": 8,
    "识别方式": 12,
    "品牌": 16,
    "主判断规则": 36,
    "二次判断/数据源": 24,
    "自动识别条件": 24,
    "冲突处理": 18,
    # 运营字段字典
    "序号": 6,
    "运营平台字段": 18,
    "业务含义": 22,
    "司机常用语言/替换词": 24,
    "Agent用途": 16,
    "是否允许对客展示": 16,
    "使用说明": 28,
    # 使用说明
    "业务": 14,
    "使用说明": 90,
}

DEFAULT_WIDTH = 16


def style_sheet(ws):
    """给单个sheet套用样式"""
    if ws.max_row == 0 or ws.max_column == 0:
        return

    # 1. 表头样式 (第1行)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 34

    # 2. 数据行样式
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

    # 3. 列宽 (按列名匹配, 未命中给默认值)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        col_name = str(ws.cell(1, col).value or "")
        width = COLUMN_WIDTHS.get(col_name, DEFAULT_WIDTH)
        ws.column_dimensions[letter].width = width

    # 4. 冻结首行
    ws.freeze_panes = "A2"

    # 5. 自动筛选
    ws.auto_filter.ref = ws.dimensions


def main():
    print(f"📂 打开: {FILE}")
    wb = load_workbook(FILE)

    for ws in wb.worksheets:
        print(f"🎨 样式化: {ws.title} ({ws.max_row}行 × {ws.max_column}列)")
        style_sheet(ws)

    wb.save(FILE)
    print(f"\n✅ 样式优化完成, 已保存: {FILE}")


if __name__ == "__main__":
    main()
