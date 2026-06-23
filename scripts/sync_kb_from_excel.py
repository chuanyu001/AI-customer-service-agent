# 全量同步润色版Excel到数据库 (以Excel为准) + 优化
# 1. Excel: 编号前缀 KB→JL; 「来源备注」→「答复策略」; 状态全部→已发布; 删除「转人工条件」列(移系统配置)
# 2. 数据库: 编号 KB→JL + 全量字段同步 + 转人工条件清空 + 19条未发布→published
# 3. 系统配置: 写入通用转人工规则
#
# 用法: python scripts/sync_kb_from_excel.py
# 注意: 运行前请关闭Excel中打开的该文件

import sys
import os
import asyncio

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from app.core.database import get_session_factory
from app.models import KnowledgeAnswer, SystemConfig
from sqlalchemy import select

EXCEL_PATH = r"D:\wuchu\Desktop\personalfiles\实习\kefu-agent\agent知识库汇总版6.22\行车记录仪Agent知识库_更新版622_润色版.xlsx"
SHEET_NAME = "01_知识问答库"

# 通用转人工规则 (从知识表移到系统配置)
TRANSFER_RULE = "知识未命中、答案低置信度、涉及费用/投诉/业务处理，或用户明确要求人工时转人工"


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("nan", "None", "NaT") else s


def to_bool(v):
    return clean(v) in ("是", "1", "yes", "true", "True")


def update_excel():
    """优化Excel: 编号KB→JL, 来源备注→答复策略, 状态全发布, 删转人工条件列"""
    print("📝 优化Excel...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]

    # 1. 编号 KB→JL
    code_col = header.index("知识编号") + 1
    n_code = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=code_col)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("KB"):
            cell.value = "JL" + cell.value[2:]
            n_code += 1
    print(f"  ✅ 编号 KB→JL: {n_code} 个")

    # 2. 来源备注 → 答复策略 (改表头)
    if "来源备注" in header:
        src_col = header.index("来源备注") + 1
        ws.cell(row=1, column=src_col).value = "答复策略"
        print(f"  ✅ 「来源备注」→「答复策略」")

    # 3. 状态全部→已发布
    status_col = header.index("状态") + 1
    n_status = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col)
        if cell.value and cell.value != "已发布":
            cell.value = "已发布"
            n_status += 1
    print(f"  ✅ 状态→已发布: {n_status} 条")

    # 4. 删除「转人工条件」列 (移到系统配置)
    if "转人工条件" in header:
        tc_col = header.index("转人工条件") + 1
        ws.delete_cols(tc_col)
        print(f"  ✅ 删除「转人工条件」列(规则移至系统配置)")

    wb.save(EXCEL_PATH)
    print("  ✅ Excel已保存\n")


async def sync_database():
    """数据库全量同步(以优化后的Excel为准)"""
    print("📦 读取Excel并同步数据库...")
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
    print(f"   {len(df)} 条知识\n")

    async with get_session_factory()() as db:
        rows = (await db.execute(
            select(KnowledgeAnswer).where(KnowledgeAnswer.business_area == "dashcam")
        )).scalars().all()
        db_map = {r.knowledge_code: r for r in rows}

        updated = 0
        for _, row in df.iterrows():
            jl_code = clean(row.get("知识编号"))
            if not jl_code.startswith("JL"):
                continue
            old_code = "KB" + jl_code[2:]
            ka = db_map.get(old_code)
            if not ka:
                continue

            ka.knowledge_code = jl_code
            ka.standard_question = clean(row.get("标准问题")) or ka.standard_question
            ka.standard_answer = clean(row.get("标准答案")) or ka.standard_answer
            ka.polished_answer = clean(row.get("润色答案")) or None
            ka.category_l1 = clean(row.get("产品模块")) or ka.category_l1
            ka.category_l2 = clean(row.get("知识类型")) or ka.category_l2

            mfr = clean(row.get("厂家"))
            ka.manufacturer = None if mfr in ("", "通用") else mfr

            ka.need_brand = to_bool(row.get("是否需要识别品牌"))
            ka.need_attachment = to_bool(row.get("是否需要附件"))

            # 转人工条件: 清空 (规则移到系统配置)
            ka.transfer_condition = ""

            # 状态: 全部 published
            ka.status = "published"

            # 答复策略 (原来源备注) → source_file字段复用
            ka.source_file = clean(row.get("答复策略")) or None

            updated += 1

        await db.commit()
    print(f"  ✅ 数据库同步: {updated} 条\n")

    # 写入通用转人工规则到系统配置
    async with get_session_factory()() as db:
        existing = await db.get(SystemConfig, "transfer_rule")
        if existing:
            existing.config_value = TRANSFER_RULE
        else:
            db.add(SystemConfig(
                config_key="transfer_rule",
                config_value=TRANSFER_RULE,
                config_type="string",
                description="通用转人工规则(从知识表迁移)",
            ))
        await db.commit()
    print(f"  ✅ 通用转人工规则已写入系统配置")


async def main():
    print(f"📂 Excel: {EXCEL_PATH}\n")
    update_excel()
    await sync_database()
    print("\n✅ 全部完成!")
    print("   优化项:")
    print("   - 编号 KB→JL (Excel+DB)")
    print("   - 「来源备注」→「答复策略」(Excel表头)")
    print("   - 19条未发布→已发布")
    print("   - 「转人工条件」移至系统配置(知识表清空)")
    print("   - 润色答案导入 polished_answer")
    print("   - 是否需要附件 10条 True→False")


if __name__ == "__main__":
    asyncio.run(main())
